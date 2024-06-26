from openpyxl import load_workbook

class ExcelHelper:
    def __init__(self, file_name):
        self.file_name = file_name

    def read_data(self, sheet_name):
        workbook = load_workbook(filename=self.file_name)
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data
